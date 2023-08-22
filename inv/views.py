
import datetime
from django.contrib.auth.models import Group


from django.shortcuts import render, redirect
from django.views import generic
from django.urls import reverse_lazy
from django.http import HttpResponseRedirect, HttpResponse

from django.contrib import messages

from django.contrib.auth.decorators import login_required, permission_required

from django.contrib.messages.views import SuccessMessageMixin

from bases.views import SinPrivilegios

from .models import Proyecto, ProyectoDetalle, Pnd, Departamento,Comunidad, Municipio, Pueblo, Poblacion

from .forms import ProyectoForm, ProyectoDetalleForm, PndForm, DepartamentoForm, MunicipioForm, PuebloForm, PoblacionForm, \
    ComunidadForm

from openpyxl import Workbook

from django.db.models import F, Func

def es_oficina_o_superusuario(user):
    return user.is_superuser or user.groups.filter(name='OFICINA').exists()

class ProyectoView(SinPrivilegios, generic.ListView):
    permission_required = 'inv.view_proyecto'
    model = Proyecto
    template_name = "inv/proyecto_list.html"
    context_object_name = "obj"

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()

        if es_oficina_o_superusuario(user):
            # Si es superusuario o pertenece al grupo OFICINA, mostrar todos los registros
            return qs

        # Filtrar los registros que el usuario ha creado
        qs = qs.filter(uc=user)

        qs = qs.annotate(lower_usuario=Func(F('uc__username'), function='LOWER'))
        qs = qs.order_by('lower_usuario')

        return qs



class ProyectoNew(SuccessMessageMixin, SinPrivilegios,generic.CreateView):
   
    permission_required = 'inv.add_proyecto'
    model = Proyecto
    template_name = "inv/proyecto_form.html"
    context_object_name = "obj"
    form_class = ProyectoForm
    success_url = reverse_lazy('inv:proyecto_list')
    success_message = "Prgrama creado con Exito"

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)


class ProyectoEdit(SuccessMessageMixin,SinPrivilegios, generic.UpdateView):
    permission_required = 'inv.change_proyecto'
    model = Proyecto
    template_name = "inv/proyecto_form.html"
    context_object_name = "obj"
    form_class = ProyectoForm
    success_url = reverse_lazy('inv:proyecto_list')
    login_url = "bases:login"
    success_message = "Programa actualizado con Exito"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)


class ProyectoDel(SinPrivilegios, generic.DeleteView):
    permission_required = 'inv.delete_proyecto'
    model = Proyecto
    template_name = 'inv/catalogos_del.html'
    context_object_name = 'obj'
    success_url = reverse_lazy("inv:proyecto_list")


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.urls import reverse_lazy
from django.http import HttpResponseRedirect
from django.contrib import messages
from .models import Proyecto, ProyectoDetalle, Departamento, Municipio, Comunidad, Pueblo, Poblacion
from .forms import ProyectoForm

@login_required(login_url='/login/')
@permission_required('inv.change_proyectodetalle', login_url='bases:sin_privilegios')
def programa(request, programa_id=None):
    template_name = "inv/programas.html"
    departamento = Departamento.objects.filter(estado=True)
    municipio = Municipio.objects.filter(estado=True)
    comunidad = Comunidad.objects.filter(estado=True)
    pueblo = Pueblo.objects.filter(estado=True)
    poblacion = Poblacion.objects.filter(estado=True)

    form_programas = {}
    contexto = {}

    if request.method == 'GET':
        form_programas = ProyectoForm()
        enc = Proyecto.objects.filter(pk=programa_id).first()

        if programa_id:
            if not enc:
                messages.error(request, 'Proyecto no existe')
                return redirect("inv:proyecto_list")

            usr = request.user
            if not es_oficina_o_superusuario(usr):
                if enc.uc != usr:
                    messages.error(request, 'Proyecto no fue creado por usted')
                    return redirect("inv:proyecto_list")

        if enc:
            det = ProyectoDetalle.objects.filter(proyecto=enc)
            e = {
                'descripcion': enc.descripcion,
                'fechaInicio': enc.fechaInicio,
                'fechaFin': enc.fechaFin,
                'donante': enc.donante,
            }

            form_programas = ProyectoForm(e)
        else:
            det = None

        contexto = {
            'encabezado': enc,
            'detalle': det,
            'form_enc': form_programas,
            'departamentos': departamento,
            'municipios': municipio,
            'comunidades': comunidad,
            'pueblos': pueblo,
            'poblaciones': poblacion
        }

    if request.method == 'POST':
        descripcion = request.POST.get("descripcion")
        periodo = request.POST.get("periodo")
        donante = request.POST.get("donante")

        if not programa_id:
            enc = Proyecto(
                descripcion=descripcion,
                periodo=periodo,
                donante=donante,
                uc=request.user
            )
            if enc:
                enc.save()
                programa_id = enc.id
        else:
            enc = Proyecto.objects.filter(pk=programa_id).first()
            if enc:
                enc.descripcion = descripcion
                enc.periodo = periodo
                enc.donante = donante
                enc.um = request.user.id
                enc.save()

        if not programa_id:
            return redirect("inv:proyecto_list")

        descripcion = request.POST.get("id_descripcion_detalle")
        objetivo = request.POST.get("id_objetivo_detalle")
        cantidad_beneficiados = request.POST.get("id_cantidad_beneficiados_detalle")
        departamento = request.POST.get("id_departamento")
        municipio = request.POST.get("id_municipio")
        comunidad_ids = request.POST.getlist("id_comunidad")
        pueblo = request.POST.get("id_pueblo_detalle")
        poblacion = request.POST.get("id_poblacion_detalle")

        det = ProyectoDetalle(
            proyecto=enc,
            descripcion=descripcion,
            objetivo=objetivo,
            cantidad_beneficiados=cantidad_beneficiados,
            departamento_id=departamento,
            municipio_id=municipio,
            pueblo_id=pueblo,
            poblacion_id=poblacion,
            uc=request.user
        )

        if det:
            det.save()
            if comunidad_ids:
                for comunidad_id in comunidad_ids:
                    comunidad_obj = Comunidad.objects.get(pk=comunidad_id)
                    det.comunidad.add(comunidad_obj)

        return redirect("inv:programas_edit", programa_id=programa_id)

    return render(request, template_name, contexto)

def es_oficina_o_superusuario(user):
    return user.is_superuser or user.groups.filter(name='OFICINA').exists()






    


from django.urls import reverse
from django.http import HttpResponseRedirect

from django.views.generic import UpdateView

class ProyectoDetalleEdit(SuccessMessageMixin, SinPrivilegios, UpdateView):
    permission_required = 'inv.change_proyectodetalle'
    model = ProyectoDetalle
    template_name = "inv/programa_form.html"
    context_object_name = "obj"
    form_class = ProyectoDetalleForm
    login_url = "bases:login"
    success_message = "Programa actualizado con éxito"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        self.success_url = self.request.POST.get('previous_page')
        return super().form_valid(form)

    def dispatch(self, request, *args, **kwargs):
        pk = self.kwargs.get('pk')
        enc = ProyectoDetalle.objects.filter(pk=pk).first()

        if not enc:
            messages.error(request, 'Detalle de proyecto no existe')
            return HttpResponseRedirect(reverse("inv:proyecto_list"))

        usr = request.user
        if not es_oficina_o_superusuario(usr) and enc.proyecto.uc != usr:
            messages.error(request, 'No tiene permiso para modificar este detalle de proyecto')
            return HttpResponseRedirect(reverse("inv:proyecto_list"))

        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(ProyectoDetalleEdit, self).get_context_data(**kwargs)
        context["departamentos"] = Departamento.objects.all()
        context["municipios"] = Municipio.objects.all()
        context["comunidades"] = Comunidad.objects.all()
        return context

def es_oficina_o_superusuario(user):
    return user.is_superuser or user.groups.filter(name='OFICINA').exists()



    
class ProyectoDetalleDel(SinPrivilegios, generic.DeleteView):
    permission_required = 'inv.delete_proyectodetalle'
    model = ProyectoDetalle
    template_name = 'inv/proyecto_det_del.html'
    context_object_name = 'obj'  

    def get_success_url(self):
        programa_id=self.kwargs['programa_id']
        return reverse_lazy('inv:programas_edit', kwargs={'programa_id':programa_id})

class PndNew(SuccessMessageMixin, SinPrivilegios,generic.CreateView):
   
    permission_required = 'inv.add_pnd'
    model = Pnd
    template_name = "inv/pnd_form.html"
    context_object_name = "obj"
    form_class = PndForm
    success_url = reverse_lazy('inv:pnd_list')
    success_message = "Prioridad Creada con Exito"

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)




class PndView(SinPrivilegios,generic.ListView):
    permission_required = 'inv.view_pnd'
    model = Pnd
    template_name  = "inv/pnd_list.html"
    context_object_name = "obj"

class PndEdit(SuccessMessageMixin,SinPrivilegios, generic.UpdateView):
    permission_required = 'inv.change_pnd'
    model = Pnd
    template_name = "inv/pnd_form.html"
    context_object_name = "obj"
    form_class = PndForm
    success_url = reverse_lazy('inv:pnd_list')
    login_url = "bases:login"
    success_message = "Prioridad actualizada con Exito"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)


class PndDel(SinPrivilegios, generic.DeleteView):
    permission_required = 'inv.delete_pnd'
    model = Pnd
    template_name = 'inv/catalogos_del.html'
    context_object_name = 'obj'
    success_url = reverse_lazy("inv:pnd_list")

class DepartamentoView(SinPrivilegios,generic.ListView):
    permission_required = 'inv.view_departamento'
    model = Departamento
    template_name  = "inv/departamento_list.html"
    context_object_name = "obj"

    def get_queryset(self):
        return Departamento.objects.order_by('descripcion')

class DepartamentoNew(SuccessMessageMixin, SinPrivilegios,generic.CreateView):
   
    permission_required = 'inv.add_departamento'
    model = Departamento
    template_name = "inv/departamento_form.html"
    context_object_name = "obj"
    form_class = DepartamentoForm
    success_url = reverse_lazy('inv:departamento_list')
    success_message = "Departamento Creada con Exito"

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)

class DepartamentoEdit(SuccessMessageMixin,SinPrivilegios, generic.UpdateView):
    permission_required = 'inv.change_departamento'
    model = Departamento
    template_name = "inv/departamento_form.html"
    context_object_name = "obj"
    form_class = DepartamentoForm
    success_url = reverse_lazy('inv:departamento_list')
    login_url = "bases:login"
    success_message = "Departamento actualizada con Exito"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)


class DepartamentoDel(SinPrivilegios, generic.DeleteView):
    permission_required = 'inv.delete_departamento'
    model = Departamento
    template_name = 'inv/catalogos_del.html'
    context_object_name = 'obj'
    success_url = reverse_lazy("inv:departamento_list")


class MunicipioView(SinPrivilegios,generic.ListView):
    permission_required = 'inv.view_municipio'
    model = Municipio
    template_name  = "inv/municipio_list.html"
    context_object_name = "obj"

    def get_queryset(self):
        return Municipio.objects.order_by('departamento')


class MunicipioNew(SuccessMessageMixin,SinPrivilegios, generic.CreateView):
    permission_required = 'inv.add_municipio'
    model = Municipio
    template_name = "inv/municipio_form.html"
    context_object_name = "obj"
    form_class = MunicipioForm
    success_url = reverse_lazy('inv:municipio_list')
    success_message = "municipio creado con Exito"
   

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)

class MunicipioDel(SinPrivilegios, generic.DeleteView):
    permission_required = 'inv.delete_municipio'
    model = Municipio
    template_name = 'inv/catalogos_del.html'
    context_object_name = 'obj'
    success_url = reverse_lazy("inv:municipio_list")

class MunicipioEdit(SuccessMessageMixin,SinPrivilegios, generic.UpdateView):
    permission_required = 'inv.change_municipio'
    model = Municipio
    template_name = "inv/municipio_form.html"
    context_object_name = "obj"
    form_class = MunicipioForm
    success_url = reverse_lazy('inv:municipio_list')
    login_url = "bases:login"
    success_message = "Municipio actualizada con Exito"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)


class PuebloView(SinPrivilegios,generic.ListView):
    permission_required = 'inv.view_pueblo'
    model = Pueblo
    template_name  = "inv/pueblo_list.html"
    context_object_name = "obj"


class PuebloNew(SuccessMessageMixin,SinPrivilegios, generic.CreateView):
    permission_required = 'inv.add_pueblo'
    model = Pueblo
    template_name = "inv/pueblo_form.html"
    context_object_name = "obj"
    form_class = PuebloForm
    success_url = reverse_lazy('inv:pueblo_list')
    success_message = "pueblo creado con Exito"
   

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)

class PuebloEdit(SuccessMessageMixin,SinPrivilegios, generic.UpdateView):
    permission_required = 'inv.change_pueblo'
    model = Pueblo
    template_name = "inv/pueblo_form.html"
    context_object_name = "obj"
    form_class = PuebloForm
    success_url = reverse_lazy('inv:pueblo_list')
    login_url = "bases:login"
    success_message = "Pueblo actualizada con Exito"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)

class PuebloDel(SinPrivilegios, generic.DeleteView):
    permission_required = 'inv.delete_pueblo'
    model = Pueblo
    template_name = 'inv/catalogos_del.html'
    context_object_name = 'obj'
    success_url = reverse_lazy("inv:pueblo_list")


class PoblacionView(SinPrivilegios,generic.ListView):
    permission_required = 'inv.view_poblacion'
    model = Poblacion
    template_name  = "inv/poblacion_list.html"
    context_object_name = "obj"


class PoblacionNew(SuccessMessageMixin,SinPrivilegios, generic.CreateView):
    permission_required = 'inv.add_poblacion'
    model = Poblacion
    template_name = "inv/poblacion_form.html"
    context_object_name = "obj"
    form_class = PoblacionForm
    success_url = reverse_lazy('inv:poblacion_list')
    success_message = "Población creado con Exito"
   

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)


class PoblacionEdit(SuccessMessageMixin,SinPrivilegios, generic.UpdateView):
    permission_required = 'inv.change_poblacion'
    model = Poblacion
    template_name = "inv/poblacion_form.html"
    context_object_name = "obj"
    form_class = PoblacionForm
    success_url = reverse_lazy('inv:poblacion_list')
    login_url = "bases:login"
    success_message = "Poblacion actualizada con Exito"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)

class PoblacionDel(SinPrivilegios, generic.DeleteView):
    permission_required = 'inv.delete_poblacion'
    model = Poblacion
    template_name = 'inv/catalogos_del.html'
    context_object_name = 'obj'
    success_url = reverse_lazy("inv:poblacion_list")



class ReporteComunidadExcel(SinPrivilegios, generic.TemplateView):
    permission_required = 'inv.add_proyecto'

    def get(self, request, *args, **kwargs):
        if request.user.is_superuser:
        # Si es superusuario, muestra todos los proyectos
         proyectos = Proyecto.objects.all()
        else:
        # Si no es superusuario, filtra los proyectos del usuario logueado
         proyectos = Proyecto.objects.filter(uc=request.user)
         

        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE PROGRAMAS'
        ws.merge_cells('B1:E1')

        ws['B2'] = 'ONG'
        ws['C2'] = 'PROYECTO'
        ws['D2'] = 'CONTACTO'
        ws['E2'] = 'TELEFONO'
        ws['F2'] = 'EMAIL'
        ws['G2'] = 'DONANTE'
        ws['H2'] = 'FECHA INICIO'
        ws['I2'] = 'FECHA FIN'
        ws['J2'] = 'OBJETIVO'
        ws['K2'] = 'CANTIDAD BENEFICIADOS'
        ws['L2'] = 'DEPARTAMENTO'
        ws['M2'] = 'MUNICIPIO'
        ws['N2'] = 'COMUNIDADES'
        ws['O2'] = 'PUEBLO'
        ws['P2'] = 'POBLACION'
        ws['Q2'] = 'MONTO PND 1'
        ws['R2'] = 'INDICADOR PND 1'
        ws['S2'] = 'META PND 1'
        ws['T2'] = 'MONTO PND 2'
        ws['U2'] = 'INDICADOR PND 2'
        ws['V2'] = 'META PND 2'
        ws['W2'] = 'MONTO PND 3'
        ws['X2'] = 'INDICADOR PND 3'
        ws['Y2'] = 'META PND 3'
        ws['Z2'] = 'MONTO PND 4'
        ws['AA2'] = 'INDICADOR PND 4'
        ws['AB2'] = 'META PND 4'
        ws['AC2'] = 'MONTO PND 5'
        ws['AD2'] = 'INDICADOR PND 5'
        ws['AE2'] = 'META PND 5'
        ws['AF2'] = 'MONTO PND 6'
        ws['AG2'] = 'INDICADOR PND 6'
        ws['AH2'] = 'META PND 6'
        ws['AI2'] = 'MONTO PND 7'
        ws['AJ2'] = 'INDICADOR PND 7'
        ws['AK2'] = 'META PND 7'
        ws['AL2'] = 'MONTO PND 8'
        ws['AM2'] = 'INDICADOR PND 8'
        ws['AN2'] = 'META PND 8'
        ws['AO2'] = 'MONTO PND 9'
        ws['AP2'] = 'INDICADOR PND 9'
        ws['AQ2'] = 'META PND 9'
        ws['AR2'] = 'MONTO PND 10'
        ws['AS2'] = 'INDICADOR PND 10'
        ws['AT2'] = 'META PND 10'
        ws['AU2'] = 'OTRO 1'
        ws['AV2'] = 'MONTO'
        ws['AW2'] = 'INDICADOR'
        ws['AX2'] = 'META'
        ws['AY2'] = 'OTRO 2'
        ws['AZ2'] = 'MONTO'
        ws['BA2'] = 'INDICADOR'
        ws['BB2'] = 'META'
        ws['BC2'] = 'OTRO 3'
        ws['BD2'] = 'MONTO'
        ws['BE2'] = 'INDICADOR'
        ws['BF2'] = 'META'
        # Agregar aquí más encabezados según las columnas adicionales del modelo ProyectoDetalle

        cont = 3

        for proyecto in proyectos:
            ws.cell(row=cont, column=2).value = proyecto.uc.username
            ws.cell(row=cont, column=3).value = proyecto.descripcion
            ws.cell(row=cont, column=4).value = proyecto.contacto
            ws.cell(row=cont, column=5).value = proyecto.telefono
            ws.cell(row=cont, column=6).value = proyecto.email
            ws.cell(row=cont, column=7).value = proyecto.donante
            ws.cell(row=cont, column=8).value = proyecto.fechaInicio
            ws.cell(row=cont, column=9).value = proyecto.fechaFin
            

            detalles = ProyectoDetalle.objects.filter(proyecto=proyecto)  # Filtrar los detalles correspondientes al proyecto

            for detalle in detalles:
                ws.cell(row=cont, column=10).value = detalle.objetivo
                ws.cell(row=cont, column=11).value = detalle.cantidad_beneficiados
                ws.cell(row=cont, column=12).value = detalle.departamento.descripcion
                ws.cell(row=cont, column=13).value = detalle.municipio.descripcion
                ws.cell(row=cont, column=14).value = ', '.join([comunidad.descripcion for comunidad in detalle.comunidad.all()])
                ws.cell(row=cont, column=15).value = detalle.pueblo.descripcion
                ws.cell(row=cont, column=16).value = detalle.poblacion.descripcion
                ws.cell(row=cont, column=17).value = detalle.montoPND1
                ws.cell(row=cont, column=18).value = detalle.indicadorPND1
                ws.cell(row=cont, column=19).value = detalle.metaPND1
                ws.cell(row=cont, column=20).value = detalle.montoPND2
                ws.cell(row=cont, column=21).value = detalle.indicadorPND2
                ws.cell(row=cont, column=22).value = detalle.metaPND2
                ws.cell(row=cont, column=23).value = detalle.montoPND3
                ws.cell(row=cont, column=24).value = detalle.indicadorPND3
                ws.cell(row=cont, column=25).value = detalle.metaPND3
                ws.cell(row=cont, column=26).value = detalle.montoPND4
                ws.cell(row=cont, column=27).value = detalle.indicadorPND4
                ws.cell(row=cont, column=28).value = detalle.metaPND4
                ws.cell(row=cont, column=29).value = detalle.montoPND5
                ws.cell(row=cont, column=30).value = detalle.indicadorPND5
                ws.cell(row=cont, column=31).value = detalle.metaPND5
                ws.cell(row=cont, column=32).value = detalle.montoPND6
                ws.cell(row=cont, column=33).value = detalle.indicadorPND6
                ws.cell(row=cont, column=34).value = detalle.metaPND6
                ws.cell(row=cont, column=35).value = detalle.montoPND7
                ws.cell(row=cont, column=36).value = detalle.indicadorPND7
                ws.cell(row=cont, column=37).value = detalle.metaPND7
                ws.cell(row=cont, column=38).value = detalle.montoPND8
                ws.cell(row=cont, column=39).value = detalle.indicadorPND8
                ws.cell(row=cont, column=40).value = detalle.metaPND8
                ws.cell(row=cont, column=41).value = detalle.montoPND9
                ws.cell(row=cont, column=42).value = detalle.indicadorPND9
                ws.cell(row=cont, column=43).value = detalle.metaPND9
                ws.cell(row=cont, column=44).value = detalle.montoPND10
                ws.cell(row=cont, column=45).value = detalle.indicadorPND10
                ws.cell(row=cont, column=46).value = detalle.metaPND10
                ws.cell(row=cont, column=47).value = detalle.PNDOtro1
                ws.cell(row=cont, column=48).value = detalle.montoOtro1
                ws.cell(row=cont, column=49).value = detalle.indicadorOtro1
                ws.cell(row=cont, column=50).value = detalle.metaOtro1
                ws.cell(row=cont, column=51).value = detalle.PNDOtro2
                ws.cell(row=cont, column=52).value = detalle.montoOtro2
                ws.cell(row=cont, column=53).value = detalle.indicadorOtro2
                ws.cell(row=cont, column=54).value = detalle.metaOtro2
                ws.cell(row=cont, column=55).value = detalle.PNDOtro3
                ws.cell(row=cont, column=56).value = detalle.montoOtro3
                ws.cell(row=cont, column=57).value = detalle.indicadorOtro3
                ws.cell(row=cont, column=58).value = detalle.metaOtro3
                # Agregar aquí más celdas según las columnas adicionales del modelo ProyectoDetalle

                cont += 1

        # Obtener la fecha actual
        fecha_actual = datetime.date.today()

            # Formatear la fecha en el formato deseado (por ejemplo, "YYYY-MM-DD")
        nombre_archivo = fecha_actual.strftime("%Y-%m-%d") + ".xlsx"

            # Resto del código
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response



class ComunidadView(SinPrivilegios,generic.ListView):
    permission_required = 'inv.view_comunidad'
    model = Comunidad
    template_name  = "inv/comunidad_list.html"
    context_object_name = "obj"

    def get_queryset(self):
        return Comunidad.objects.order_by('municipio')
       
   
class ComunidadNew(SuccessMessageMixin, SinPrivilegios,generic.CreateView):
   
    permission_required = 'inv.add_comunidad'
    model = Comunidad
    template_name = "inv/comunidad_form.html"
    context_object_name = "obj"
    form_class = ComunidadForm
    success_url = reverse_lazy('inv:comunidad_list')
    success_message = "Comunidad Creada con Exito"

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super(ComunidadNew, self).get_context_data(**kwargs)
        context["departamentos"] = Departamento.objects.all()
        context["municipios"] = Municipio.objects.all()
        #print(context)
        return context


class ComunidadEdit(SuccessMessageMixin,SinPrivilegios, generic.UpdateView):
    permission_required = 'inv.change_comunidad'
    model = Comunidad
    template_name = "inv/comunidad_form.html"
    context_object_name = "obj"
    form_class = ComunidadForm
    success_url = reverse_lazy('inv:comunidad_list')
    login_url = "bases:login"
    success_message = "Comunidad actualizada con Exito"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)
    def get_context_data(self, **kwargs):
        pk = self.kwargs.get('pk')
        context = super(ComunidadEdit, self).get_context_data(**kwargs)
        context["departamentos"] = Departamento.objects.all()
        context["municipios"] = Municipio.objects.all()
        context["obj"] = Comunidad.objects.filter(pk=pk).first()
        return context


class ComunidadDel(SinPrivilegios, generic.DeleteView):
    permission_required = 'inv.delete_comunidad'
    model = Comunidad
    template_name = 'inv/catalogos_del.html'
    context_object_name = 'obj'
    success_url = reverse_lazy("inv:comunidad_list")




@login_required(login_url='/login/')
@permission_required('inv.view_comunidad', login_url='bases:sin_privilegios')
def dashboard(request):
    template_name = "inv/dashboard.html"

    detalles = ProyectoDetalle.objects.all()
    proyectos = Proyecto.objects.count
    programas = ProyectoDetalle.objects.count

    grupo_cooperantes = Group.objects.get(name='COOPERANTES')
    cantidad_cooperantes = grupo_cooperantes.user_set.count()

    grupo_instituciones = Group.objects.get(name='INSTITUCIONES')
    cantidad_instituciones = grupo_instituciones.user_set.count()

    
    # Calcula las sumas de montosPND del 1 al 10 y montosOtro del 1 al 3
    sumas_montos_pnd = [sum(float(getattr(detalle, f'montoPND{i}', 0)) for detalle in detalles) for i in range(1, 11)]
    sumas_montos_otro = [sum(float(getattr(detalle, f'montoOtro{i}', 0)) for detalle in detalles) for i in range(1, 4)]
    

    context = {
        'sumas_montos_pnd': sumas_montos_pnd,
        'sumas_montos_otro': sumas_montos_otro,
        'proyectos': proyectos,
        'programas': programas,
        'cantidad_cooperantes':cantidad_cooperantes,
        'cantidad_instituciones':cantidad_instituciones
    }

    return render(request, template_name, context)


class DirectorioView(SinPrivilegios,generic.ListView):
    permission_required = 'inv.view_comunidad'
    model = Proyecto
    template_name  = "inv/directorio_list.html"
    context_object_name = "obj"




   
   